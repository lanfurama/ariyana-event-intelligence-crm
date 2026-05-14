"""Build Excel from scraped ICCA data."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
items = json.load(open(ROOT / 'all_items.json'))
orgs = json.load(open(ROOT / 'all_orgs.json'))

wb = Workbook()

# --- Sheet 1: Series with org details denormalized ---
ws1 = wb.active
ws1.title = 'Series'
headers1 = [
    'page', 'series_id', 'series_name', 'attendance', 'series_url',
    'org_id', 'org_name', 'acode', 'org_url',
    'address', 'phone', 'fax', 'website', 'email',
    'twitter', 'facebook', 'linkedin', 'instagram',
    'notes',
    'contacts_count', 'contacts_summary'
]
ws1.append(headers1)

for it in items:
    org = orgs.get(it['org_id']) if it['org_id'] else None
    if org and 'error' not in org:
        social = org.get('social', {})
        contacts = org.get('contacts', [])
        contact_summary = '; '.join(
            f"{c['name']}" + (f" ({c['title']})" if c['title'] else '') +
            (f" <{c['email']}>" if c['email'] else '') +
            (f" {c['phone']}" if c['phone'] else '')
            for c in contacts
        )
        row = [
            it['source_page'], it['series_id'], it['series_name'], it['attendance'], it['series_url'],
            it['org_id'], org.get('name', it['org_name']), org.get('acode', ''), it['org_url'],
            org.get('address', ''), org.get('phone', ''), org.get('fax', ''),
            org.get('website', ''), org.get('email', ''),
            social.get('twitter', ''), social.get('facebook', ''),
            social.get('linkedin', ''), social.get('instagram', ''),
            org.get('notes', ''),
            len(contacts), contact_summary
        ]
    else:
        row = [
            it['source_page'], it['series_id'], it['series_name'], it['attendance'], it['series_url'],
            it['org_id'], it['org_name'], '', it['org_url'],
            '', '', '', '', '',
            '', '', '', '',
            '',
            0, ''
        ]
    ws1.append(row)

# --- Sheet 2: Organizations (one row per unique org) ---
ws2 = wb.create_sheet('Organizations')
headers2 = [
    'org_id', 'name', 'acode', 'org_url',
    'address', 'phone', 'fax', 'website', 'email',
    'twitter', 'facebook', 'linkedin', 'instagram',
    'notes', 'contacts_count'
]
ws2.append(headers2)
for oid, org in sorted(orgs.items(), key=lambda x: int(x[0])):
    if 'error' in org:
        ws2.append([oid] + [f"ERROR: {org['error']}"] + [''] * 13)
        continue
    social = org.get('social', {})
    ws2.append([
        oid, org.get('name', ''), org.get('acode', ''),
        f"https://icca.amd.simpleviewinc.com/ViewOrganization.aspx?ID={oid}",
        org.get('address', ''), org.get('phone', ''), org.get('fax', ''),
        org.get('website', ''), org.get('email', ''),
        social.get('twitter', ''), social.get('facebook', ''),
        social.get('linkedin', ''), social.get('instagram', ''),
        org.get('notes', ''), len(org.get('contacts', []))
    ])

# --- Sheet 3: Contacts (one row per contact) ---
ws3 = wb.create_sheet('Contacts')
headers3 = ['org_id', 'org_name', 'contact_id', 'name', 'title', 'email', 'phone', 'last_updated']
ws3.append(headers3)
for oid, org in sorted(orgs.items(), key=lambda x: int(x[0])):
    if 'error' in org:
        continue
    org_name = org.get('name', '')
    for c in org.get('contacts', []):
        ws3.append([
            oid, org_name, c.get('id', ''),
            c.get('name', ''), c.get('title', ''),
            c.get('email', ''), c.get('phone', ''),
            c.get('last_updated', '')
        ])

# --- Style: bold header, autosize, freeze first row ---
def style_sheet(ws):
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    # Auto width (cap at 60)
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            v_str = str(v)
            # Take first line for width calc (multiline addresses)
            v_str = v_str.split('\n')[0]
            max_len = max(max_len, len(v_str))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

for ws in (ws1, ws2, ws3):
    style_sheet(ws)

out = ROOT / 'icca_vietnam_meetings.xlsx'
wb.save(out)
print(f'Saved {out}')
print(f'Sheet "Series": {ws1.max_row - 1} rows')
print(f'Sheet "Organizations": {ws2.max_row - 1} rows')
print(f'Sheet "Contacts": {ws3.max_row - 1} rows')
